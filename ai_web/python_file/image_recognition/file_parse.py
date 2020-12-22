import openpyxl

class Getdata():
    def __init__(self):
        self.fan_runtime = 0
        self.file_name = 0
        self.iv = []


class PmiParse():
    def __init__(self,fileIn):
        self.getdata = Getdata()
        self.getdata.file_name = fileIn
        self.FileParse()

    def FileParse(self):
        data = openpyxl.load_workbook(self.getdata.file_name)
        Sheet1 = data.get_sheet_by_name('Sheet1')
        row = 4
        column = 2
        while Sheet1.cell(row=row,column=column).value != None:
            iv = []
            i = 2
            while Sheet1.cell(row=row,column=column+i).value != None:
                iv_index=[]
                iv_index.append(Sheet1.cell(row=row,column=column+i).value)
                iv_index.append(Sheet1.cell(row=row+1,column=column+i).value)
                iv.append(iv_index)
                i=i+1
            self.getdata.iv.append(iv)
            row = row + 2
